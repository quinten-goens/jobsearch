#!/usr/bin/env python3
"""
check_links.py - verify every Direct link in the Brussels job search workbook.

Run locally (NOT in Claude's sandbox - its proxy blocks all these domains).

    pip install openpyxl requests
    python check_links.py Sarah_Pernet_Brussels_Job_Search.xlsx

Writes Sarah_Pernet_Brussels_Job_Search_CHECKED.xlsx with three new columns:
    S: HTTP status
    T: Final URL after redirects
    U: Verdict  (OK / REDIRECTED / DEAD / BLOCKED / TIMEOUT)

Rows verdicted DEAD are highlighted red; REDIRECTED are highlighted amber with the
new location in column T so you can paste it back over the old link.

IMPORTANT CAVEAT: a 403 does not mean the page is gone. Plenty of sites
(Cloudflare, LinkedIn, some .gov domains) reject scripted requests regardless of
whether the URL is valid. Those are marked BLOCKED, not DEAD - check them by hand
or just use the Search link column.
"""
import sys, time
from concurrent.futures import ThreadPoolExecutor
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

URL_COL, CONF_COL = 16, 17          # P, Q
STATUS_COL, FINAL_COL, VERDICT_COL = 19, 20, 21   # S, T, U

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")
HEADERS = {"User-Agent": UA, "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
           "Accept-Language": "en-GB,en;q=0.9,fr;q=0.8"}

RED   = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREEN = PatternFill("solid", fgColor="C6EFCE")


def check(url: str):
    """Return (status, final_url, verdict)."""
    if not url:
        return ("", "", "")
    for method in ("head", "get"):
        try:
            r = requests.request(
                method, url, headers=HEADERS, timeout=15,
                allow_redirects=True, stream=(method == "get"))
            code = r.status_code
            final = r.url
            if method == "get":
                r.close()
            # Some servers refuse HEAD but serve GET fine -> retry as GET
            if method == "head" and code in (403, 405, 501):
                continue
            if code in (401, 403, 429):
                return (code, final, "BLOCKED")
            if code >= 500:
                return (code, final, "BLOCKED")   # server-side, not necessarily dead
            if code >= 400:
                return (code, final, "DEAD")
            if final.rstrip("/") != url.rstrip("/"):
                return (code, final, "REDIRECTED")
            return (code, final, "OK")
        except requests.exceptions.Timeout:
            if method == "get":
                return ("", "", "TIMEOUT")
        except requests.exceptions.SSLError as e:
            return ("", "", f"SSL: {str(e)[:60]}")
        except requests.exceptions.RequestException as e:
            if method == "get":
                return ("", "", f"DEAD ({type(e).__name__})")
    return ("", "", "DEAD")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "Sarah_Pernet_Brussels_Job_Search.xlsx"
    wb = load_workbook(path)
    ws = wb["Target List"]

    rows = [(r, ws.cell(r, URL_COL).value) for r in range(2, ws.max_row + 1)]
    rows = [(r, u) for r, u in rows if u]
    print(f"Checking {len(rows)} URLs with 12 workers...\n")

    ws.cell(1, STATUS_COL, "HTTP status").font = Font(bold=True)
    ws.cell(1, FINAL_COL, "Final URL").font = Font(bold=True)
    ws.cell(1, VERDICT_COL, "Verdict").font = Font(bold=True)

    t0 = time.time()
    with ThreadPoolExecutor(max_workers=12) as pool:
        results = list(pool.map(lambda x: check(x[1]), rows))

    tally = {}
    for (r, url), (code, final, verdict) in zip(rows, results):
        ws.cell(r, STATUS_COL, code)
        ws.cell(r, FINAL_COL, final if verdict == "REDIRECTED" else "")
        c = ws.cell(r, VERDICT_COL, verdict)
        key = verdict.split()[0].split(":")[0]
        tally[key] = tally.get(key, 0) + 1
        if verdict.startswith("DEAD"):
            c.fill = RED
            ws.cell(r, URL_COL).fill = RED
        elif verdict == "REDIRECTED":
            c.fill = AMBER
            ws.cell(r, FINAL_COL).fill = AMBER
        elif verdict == "OK":
            c.fill = GREEN
        print(f"{verdict:<12} {str(code):<5} {url}")

    for w, col in ((12, "S"), (46, "T"), (14, "U")):
        ws.column_dimensions[col].width = w

    out = path.replace(".xlsx", "_CHECKED.xlsx")
    wb.save(out)

    print("\n" + "=" * 60)
    for k, v in sorted(tally.items(), key=lambda kv: -kv[1]):
        print(f"  {k:<12} {v}")
    print(f"\n  {time.time()-t0:.0f}s elapsed")
    print(f"  Written: {out}")
    print("\n  RED rows    = replace the link (use the Search link column).")
    print("  AMBER rows  = site moved; paste column T over column P.")
    print("  BLOCKED     = bot protection, not necessarily broken. Check by hand.")


if __name__ == "__main__":
    main()

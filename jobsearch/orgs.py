"""Load the Target List sheet into orgs.json.

The workbook stays Sarah's manual tracker; we only ever read it.
"""
import json

from openpyxl import load_workbook

from .config import ORGS_JSON, WORKBOOK

COLS = {
    "id": 1,
    "organisation": 2,
    "category": 3,
    "type": 4,
    "base": 5,
    "target_roles": 6,
    "why_fits": 7,
    "language_edge": 8,
    "eu_nationality": 9,
    "priority": 10,
    "existing_url": 16,
    "existing_confidence": 17,
}


def load_orgs() -> list[dict]:
    ws = load_workbook(WORKBOOK, read_only=True, data_only=True)["Target List"]
    orgs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[COLS["organisation"] - 1]:
            continue
        org = {k: row[i - 1] for k, i in COLS.items()}
        org["organisation"] = str(org["organisation"]).strip()
        orgs.append(org)
    return orgs


def main() -> None:
    orgs = load_orgs()
    ORGS_JSON.write_text(json.dumps(orgs, indent=2, ensure_ascii=False, default=str))
    print(f"Wrote {len(orgs)} orgs -> {ORGS_JSON}")


if __name__ == "__main__":
    main()
